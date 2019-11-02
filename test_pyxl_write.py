import datetime
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = datetime.datetime(2010,7,21)
ws['B1'] = datetime.date(2019,4,10)
c1 = ws['B1']
print(type(c1.value))
print(c1.value)
print(ws['A1'].number_format)
wb.save('date.xlsx')